# encoding: utf-8
import json
from openpyxl import load_workbook
from openpyxl.cell import Cell
wb = load_workbook('nor.xlsx')
li_na ={}
cur_sheet = wb.get_sheet_by_name("网站货币")
for row in range(2, cur_sheet.max_row + 1):
    marketCode = cur_sheet['A' + str(row)].value
    nation = cur_sheet['B' + str(row)].value
    currency = cur_sheet['C' + str(row)].value
    li_na[nation] = {"marketCode" : marketCode, 'currency': currency}
    print(marketCode, nation, currency)

ws = wb.get_sheet_by_name('航线始发国')
data_dict = {}
# print(dir(ws))
for row in range(2, ws.max_row +1):
    dep = ws['A' + str(row)].value
    dt = ws['B' + str(row)].value
    nation = ws['C' + str(row)].value
    data_dict.setdefault(dep, li_na.get(nation))
    # print(dep, dt, nation)
print(data_dict)
f = open("depport_currency.json", "w")
json.dump(data_dict, f, indent=4)



